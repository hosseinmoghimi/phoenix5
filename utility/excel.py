from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from phoenix.server_settings import STATIC_ROOT


def get_excel_report():
    wb = Workbook()

    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ws1.append(range(600))

    ws2 = wb.create_sheet(title="Pi")

    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)

    return wb





from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.http import HttpResponse
from .calendar import PersianCalendar
from datetime import datetime
from openpyxl import load_workbook
from core.settings import *
import os
def get_style(size=11,font_name='Calibri',bold=False,locked=True,color='FF000000',start_color='FFFFFF',end_color='FF000000',*args, **kwargs):
    font = Font(name=font_name,
                    size=size,
                    bold=bold,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color=color)
    fill = PatternFill(fill_type=None,
                    start_color=start_color,
                    end_color=end_color)
    border_left=None
    if 'border_left' in kwargs or 'bordered' in kwargs:
        border_left=Side(border_style='thin',color='000000')
    border_right=None
    if 'border_right' in kwargs or 'bordered' in kwargs:
        border_right=Side(border_style='thin',color='000000')
    border_top=None
    if 'border_top' in kwargs or 'bordered' in kwargs:
        border_top=Side(border_style='thin',color='000000')
    border_bottom=None
    if 'border_bottom' in kwargs or 'bordered' in kwargs:
        border_bottom=Side(border_style='thin',color='000000')

    border = Border(left=border_left,right=border_right,top=border_top,bottom=border_bottom,
                    diagonal=Side(border_style=None,
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                color='FF000000'),
                    vertical=Side(border_style=None,
                                color='FF000000'),
                    horizontal=Side(border_style=None,
                                color='FF000000')
                )
    alignment=Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)
    number_format = 'General'
    protection = Protection(locked=locked,
                            hidden=False)
    return {'font':font,
        'fill':fill,
        'border':border,
        'alignment':alignment,
        'number_format':number_format,
        'protection':protection,
    }
class ReportWorkBook:
    
    def __init__(self,origin_file_name=None,*args, **kwargs):
        self.sheets=[]
        self.origin_file_name=origin_file_name
        self.sheet_counter=0

        if self.origin_file_name is None:
            self.work_book = Workbook()
            for sheet in self.sheets:
                self.work_book.create_sheet(sheet.sheet_name)
        else:
            REPORT_ROOT=os.path.join(STATIC_ROOT,'report')
            filename =os.path.join(REPORT_ROOT,self.origin_file_name)   
            # filename =os.path.join(STATIC_ROOT,self.origin_file_name)      
            self.work_book = load_workbook(filename = filename)
        
    # def to_excel2(self):
    #     date=PersianCalendar().from_gregorian(datetime.now())
    #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    #     # response.AppendHeader("Content-Type", "application/vnd.ms-excel");
    #     response["Content-disposition"]=f"attachment; filename={self.file_name}-{date}.xlsx"
    #     work_book=Workbook()
    #     for i,sheet in enumerate(self.sheets):
    #         if i==0:
    #             worksheet=work_book.active
    #         else:
    #             worksheet=work_book.create_sheet(f'sheet#{i}')
    #         sheet.get_worksheet(worksheet,blank_sheet=True)
    #     work_book.save(response)
    #     return response
    def add_sheet(self,style=None,data=None,start_row=1,start_col=1,sheet_name='گزارش',title=None,blank_sheet=None,table_headers=None,*args, **kwargs):
        sheet=ReportSheet()
        sheet.data=data
        sheet.blank_sheet=blank_sheet
        sheet.start_row=start_row
        sheet.start_col=start_col
        if style is None:
            style=get_style()     
        sheet.style=style
        sheet.sheet_name=sheet_name
        sheet.title=title
        
        sheet.table_headers=table_headers


        self.sheet_counter+=1
        if self.origin_file_name is None:
            self.work_book.create_sheet(sheet.sheet_name)
        worksheet=self.work_book.worksheets[self.sheet_counter]
        current_row=sheet.start_row
        start_col=sheet.start_col
        
        worksheet.sheet_name = sheet.sheet_name
        worksheet.sheet_view.rightToLeft = True
        if len(sheet.data)<1:
            return None

        # Define the titles for columns

        # cell = worksheet.cell(row=current_row, column=1)
        # cell.value = self.title
        # cell.border=style['border']
        # cell.alignment=style['alignment']
        # current_row+=1
        if sheet.blank_sheet:
            column=len( sheet.data[0].keys())
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column+start_col)       
        # Assign the titles for each cell of the header
        if sheet.table_headers is not None:
            for col_num, column_title in enumerate(sheet.table_headers, 0):
                cell = worksheet.cell(row=current_row, column=col_num+start_col)
                cell.value = column_title
                if sheet.blank_sheet:
                    cell.alignment=sheet.style['alignment']
                    cell.border=sheet.style['border']
            current_row+=1
        # Iterate through all movies
        for data_item in sheet.data:
            
            col_num=0
            # Assign the data for each cell of the row 
            for cell_value in data_item:
                cell = worksheet.cell(row=current_row, column=col_num+start_col)
                cell.value = data_item[cell_value]
                cell.border=sheet.style['border']
                cell.font=sheet.style['font']
                col_num+=1
            current_row += 1


    
class ReportSheet:
    def __init__(self,*args, **kwargs):
        pass
        # self.data=data
        # self.blank_sheet=blank_sheet
        # self.start_row=start_row
        # self.start_col=start_col
        # if style is None:
        #     style=get_style()     
        # self.style=style
        # self.sheet_name=sheet_name
        # self.title=title
        
        # self.table_headers=table_headers
    
    